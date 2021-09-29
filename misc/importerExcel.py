from iptracker import Thread
import sqlite3
import calendar
from openpyxl import Workbook, load_workbook

con = sqlite3.connect("wahThreads.db")
cur = con.cursor()

wb = load_workbook(filename="q518z8.xlsx")
ws = wb.active

start = 36
end = 77

i = start 
while i <= end:
    year = "21"
    month = str(ws['B' + str(i)].value)[5:7]
    date = str(ws['B' + str(i)].value)[8:10]
    day = calendar.day_name[ws['B' + str(i)].value.weekday()][:3]

    time = str(ws['C' + str(i)].value)

    now = month + r'\/' + date + r'\/' + year + '(' + day + ')' + time

    replies = str(ws['D' + str(i)].value)

    images = str(ws['E' + str(i)].value)

    ips = str(ws['F' + str(i)].value)

    idd = str(ws['G' + str(i)].value)[38:45]

    i += 1

    cur.execute('INSERT or REPLACE INTO Threads VALUES (?, ?, ?, ?, ?, ?, ?)', 
            (int(idd),
            now, 
            "Anonymous", 
            r"\/wah\/", 
            int(replies),
            int(images),
            int(ips)))

    con.commit()
